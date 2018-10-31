import json
import collections
from openpyxl import load_workbook
from openpyxl import Workbook
import xmltodict
import sys
import copy

class SkosReaderWriter():
    def __init__(self):
        self.totalWordsChecked = 0
        self.totalPrefMatchs = 0
        self.totalAltMatchs = 0

    def readSkosJson(self,path):
        #-vars used to track words
        totalwords = 0
        totalPrefLabelUpdate = 0
        totalAltLabelUpdate = 0
        try:
            temp = collections.defaultdict(lambda: collections.defaultdict(list))
            file = open(path, "rb")
            items = json.load(file)
            for key in items.keys():
                # key = URI

                for key2 in items[key].keys():

                    #for key3 in items[key][key2]:
                     #   print("key 1 is ", key)
                     #   print("key2 is ",key2)
                     #   print("key3 is ", key3)
                    #- key2 = elements in the URI
                    #print(key2)
                    prefLabel = ""
                    altLabel = []
                    broader = []
                    #-check for prefLabel
                    if "core#prefLabel" in key2:
                        #- pref label
                        prefLabel = items[key][key2][0]['value']
                        temp[prefLabel]['source'].append(key)

                    # - skos:altLabel
                    if 'core#altLabel' in key2:
                        for altItem in items[key][key2]:
                            altLabel.append(altItem['value'])
                        temp[prefLabel]['altLabel'] = altLabel

                    if 'core#broader' in key2:
                        for broaderItem in items[key][key2]:
                            broader.append(broaderItem['value'])
                        temp[prefLabel]['broader'] = broader
            print(path, " ", len(temp.keys()))
            return temp
        except:
            print("Failed to load " + path)
            return temp


    def readSkosXml(self,path):
        temp = collections.defaultdict(lambda: collections.defaultdict(list))
        #-load the xml to dictionary
        # -read the document
        with open(path, encoding="utf8") as fd:
            doc = xmltodict.parse(fd.read())

            # -places into nice format
            #{prefLabel : {altLabel: [], description : [], notation : [], note : [], source : [], broader : []}}
        try:
            startKey = ""
            if 'skos:Concept' in doc['rdf:RDF']:
                startKey = "skos:Concept"
            elif 'rdf:Description' in doc['rdf:RDF']:
                startKey = 'rdf:Description'

            for ele in doc['rdf:RDF'][startKey]:
                #print(ele)
                prefLabel = ""
                #- Store the prefLabel
                if "skos:prefLabel" in ele.keys():
                    #-Sometimes theres two pref labels, which causes errors
                    if type(ele['skos:prefLabel']) == list:
                        for everyPref in ele['skos:prefLabel']:
                            prefLabel = everyPref['#text']
                    else:  # -only one skos:inScheme
                        prefLabel = ele['skos:prefLabel']['#text']

                    # - skos:about
                    if '@rdf:about' in ele.keys():
                        # -more than one broader
                        if type(ele['@rdf:about']) == list:
                            for everyAbout in ele['@rdf:about']:
                                temp[prefLabel]['@rdf:about'].append(everyAbout)
                        else:  # -only one broader
                            temp[prefLabel]['source'] = ele['@rdf:about']

                    #- skos:inScheme
                    if 'skos:inScheme' in ele.keys():
                        # -more than one skos:inScheme
                        if type(ele['skos:inScheme']) == list:
                            for everyScheme in ele['skos:inScheme']:
                                temp[prefLabel]['inScheme'].append(everyScheme['@rdf:resource'])
                        else:  # -only one skos:inScheme
                            temp[prefLabel]['inScheme'].append(ele['skos:inScheme']['@rdf:resource'])

                    #- skos:broader
                    if 'skos:broader' in ele.keys():
                        #-more than one broader
                        if type(ele['skos:broader']) == list:
                            for everyBroader in ele['skos:broader']:
                                temp[prefLabel]['broader'].append(everyBroader['@rdf:resource'])
                        else: #-only one broader
                            temp[prefLabel]['broader'].append(ele['skos:broader']['@rdf:resource'])

                    # - skos:altLabel
                    if 'skos:altLabel' in ele.keys():
                        #- More then one alt label
                        if type(ele['skos:altLabel']) == list:
                            for everyAlt in ele['skos:altLabel']:
                                if '#text' in everyAlt:
                                    temp[prefLabel]['altLabel'].append(everyAlt['#text'])
                        else: #- only one alt label
                            if '#text' in ele['skos:altLabel']:
                                temp[prefLabel]['altLabel'].append(ele['skos:altLabel']['#text'])
            print("success : " + path + " "+ str(len(temp)))

        except:
            print("failed : " + path, sys.exc_info()[0])

        return temp

    def containsKey(self,dictionary,key):
        if key in dictionary.keys():
            return True
        else:
            return False

    def writeToExcel(self, path, data):
        # -open/create excel book
        wb = Workbook()
        ws = wb.active
        # -sets up sheet
        ws.title = "registerinfo"
        ws = wb.create_sheet("register")
        ws = wb["register"]
        ws["A1"] = "id"
        ws["B1"] = "label"
        ws["C1"] = "altLabel"
        ws["D1"] = "description"
        ws["E1"] = "notation"
        ws["F1"] = "note"
        ws["G1"] = "source"
        ws["H1"] = "broader"
        listOfKeys = list(data.keys())
        print(data.keys())
        counter = 2
        for i in range(0, len(listOfKeys)):
            if listOfKeys[i] is not "" and listOfKeys[i] is not None:
                #-ID
                ws["A" + str(counter)] = str(counter)
                #- preflabel
                ws["B" + str(counter)] = listOfKeys[i]
                #- altLabel
                if len(data[listOfKeys[i]]['altLabel']) > 0:
                    tempString = ""
                    for altLabel in data[listOfKeys[i]]['altLabel']:
                        if altLabel is not None and altLabel != "None":
                            tempString += altLabel + "\n"
                    ws["C"+str(counter)] = tempString
                #- description

                if len(data[listOfKeys[i]]['description']) > 0:
                    tempString = ""
                    for description in data[listOfKeys[i]]['description']:
                        if description is not None:
                            tempString += description + "\n"
                    ws["D"+str(counter)] = tempString

                #-notation
                if len(data[listOfKeys[i]]['notation']) > 0:
                    tempString = ""
                    for notation in data[listOfKeys[i]]['notation']:
                        if notation is not None:
                            tempString += notation + "\n"
                    ws["E"+str(counter)] = tempString

                #-note
                if len(data[listOfKeys[i]]['note']) > 0:
                    tempString = ""
                    for note in data[listOfKeys[i]]['note']:
                        if note is not None:
                            tempString += note + "\n"
                    ws["F"+str(counter)] = tempString

                #-source
                if len(data[listOfKeys[i]]['source']) > 0:
                    if data[listOfKeys[i]]['source'][0] is not None and data[listOfKeys[i]]['source'] != "None":
                        if type(data[listOfKeys[i]]['source']) == str:
                            ws["G"+str(counter)] = data[listOfKeys[i]]['source']
                        elif type(data[listOfKeys[i]]['source']) == list:
                            ws["G" + str(counter)] = data[listOfKeys[i]]['source'][0]
                # -broader
                if len(data[listOfKeys[i]]['broader']) > 0:
                    tempString = ""
                    for broader in data[listOfKeys[i]]['broader']:
                        if broader is not None:
                            tempString += broader + "\n"
                    ws["H" + str(counter)] = tempString
                counter += 1
        wb.save(path)

    def loadFromExcel(self,path):
        #- loads the data from excel and places it into
        #- {prefLabel : {altLabel: [], description : [], notation : [], note : [], source : [], broader : []}}
        temp = collections.defaultdict(lambda: collections.defaultdict(list))
        wb = load_workbook(path,data_only=True)
        ws = wb['register']
        counter = 2
        strCounter = str(counter)
        while ws["A" + strCounter].value is not None:
            strCounter = str(counter)
            prefLabel = ws["B" + strCounter].value

            altLabel = []
            for item in str(ws["C" + strCounter].value).split("|"):
                altLabel.append(item)

            description = ws["D" + strCounter].value
            notation = ws["E" + strCounter].value
            note = ws["F" + strCounter].value
            source = ws["G" + strCounter].value
            broader = ws["H" + strCounter].value
            counter += 1

            temp[prefLabel]['altLabel'] = altLabel
            temp[prefLabel]['description'].append(description)
            temp[prefLabel]['notation'].append(notation)
            temp[prefLabel]['note'].append(note)
            temp[prefLabel]['broader'].append(broader)
            temp[prefLabel]['source'].append(source)
        return temp


    def mergeSkosDicts(self, mainKeywords, extraKeywords):
        #- get list of the mainKeys
        msKeys = list(mainKeywords.keys())
        #- lower the case for the keywords
        lowerMsKeys = list(mainKeywords.keys())

        #- do same for extra keywords
        esKeys = list(extraKeywords.keys())
        esLowerKeys = list(esKeys)

        for key in esLowerKeys:
            if type(key) == str:
                key = key.lower()


        for key in lowerMsKeys:
            if type(key) == str:
                key = key.lower()

        #-vars used to track words
        totalwords = 0
        totalPrefLabelUpdate = 0
        totalAltLabelUpdate = 0

        #-create storage for keywords
        newKeywords = copy.deepcopy(mainKeywords)
        wb = Workbook()
        totalXmlKeys = []


        #- for every key in xml Keys
        for i in range(len(extraKeywords)):
            #- check to see if theres pref label match
            if esLowerKeys[i] in lowerMsKeys:
                totalwords += 1
                #-found same pref label
                tempMsKeyIndex = lowerMsKeys.index(esLowerKeys[i])
                #-check size if more alternative words, copy over
                if len(newKeywords[msKeys[tempMsKeyIndex]]['altLabel']) <= len(extraKeywords[esKeys[i]]['altLabel']):
                    # - copy
                    print("found this ", esKeys[i])
                    print("with these alt labels ", extraKeywords[esKeys[i]])
                    newKeywords[msKeys[tempMsKeyIndex]]['altLabel'] = copy.deepcopy(extraKeywords[esKeys[i]]['altLabel'])
                    newKeywords[msKeys[tempMsKeyIndex]]['description'] = copy.deepcopy(extraKeywords[esKeys[i]]['description'])
                    newKeywords[msKeys[tempMsKeyIndex]]['notation'] = copy.deepcopy(extraKeywords[esKeys[i]]['notation'])
                    newKeywords[msKeys[tempMsKeyIndex]]['note'] = copy.deepcopy(extraKeywords[esKeys[i]]['note'])
                    newKeywords[msKeys[tempMsKeyIndex]]['broader'] = copy.deepcopy(extraKeywords[esKeys[i]]['broader'])
                    newKeywords[msKeys[tempMsKeyIndex]]['source'] = copy.deepcopy(extraKeywords[esKeys[i]]['source'])
                    print("Replaced now ", newKeywords[msKeys[lowerMsKeys.index(esLowerKeys[i])]]['altLabel'])
                    totalPrefLabelUpdate += 1

        #-check if word is in the altLabel
        #-for every key in the loaded XML dictionary
        for i in range(0,len(esLowerKeys)):
            #-check if the key is in its altLabel
            for x in range(0,len(msKeys)):
                tempList = [b.lower() for b in extraKeywords[esKeys[i]]['altLabel']]
                if lowerMsKeys[x] in tempList:
                    totalwords += 1
                    #- found pref label within altLabel
                    #xmlKeywords[xmlKeywords[i]]['altLabel']
                    #- check to see which has more alt Labels
                    if len(newKeywords[msKeys[x]]['altLabel']) <= len(extraKeywords[esKeys[i]]['altLabel']):
                        #place new one inside
                        newKeywords[esKeys[i]]['altLabel'] = copy.deepcopy(extraKeywords[esKeys[i]]['altLabel'])
                        newKeywords[esKeys[i]]['description'] = copy.deepcopy(extraKeywords[esKeys[i]]['description'])
                        newKeywords[esKeys[i]]['notation'] = copy.deepcopy(extraKeywords[esKeys[i]]['notation'])
                        newKeywords[esKeys[i]]['note'] = copy.deepcopy(extraKeywords[esKeys[i]]['note'])
                        newKeywords[esKeys[i]]['broader'] = copy.deepcopy(extraKeywords[esKeys[i]]['broader'])
                        newKeywords[esKeys[i]]['source'] = copy.deepcopy(extraKeywords[esKeys[i]]['source'])
                        #delete old one
                        del newKeywords[msKeys[x]]
                        totalAltLabelUpdate += 1

        print(newKeywords)
        print("total words that was found and checked", totalwords)
        print("total pref label accepted ", totalPrefLabelUpdate, " / ", len(msKeys))
        print("total alt label checked ", totalAltLabelUpdate, " / ", len(msKeys))
        self.totalPrefMatchs += totalPrefLabelUpdate
        self.totalAltMatchs += totalAltLabelUpdate
        self.totalWordsChecked += totalwords
        print(len(set(lowerMsKeys).difference(set(totalXmlKeys))))
        return newKeywords