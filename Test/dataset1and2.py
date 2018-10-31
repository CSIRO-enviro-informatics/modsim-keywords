from collections import defaultdict
from PDF_Keywords_Extractor import PdfExtractor
import json
import types
from openpyxl import load_workbook
from openpyxl import Workbook
import xmltodict
def fileLinkReverter(name):
    return "http://mssanz.org.au" + name.replace("{WS_FSLASH}", '/')

def writeToExcel(path, data):
    # -open/create excel book
    wb = Workbook()
    ws = wb.active
    # -sets up sheet
    ws.title = "registerinfo"
    ws = wb.create_sheet("register")
    ws = wb["registerinfo"]
    ws["A1"] = "id"
    ws["B1"] = "label"
    ws["C1"] = "altLabel"
    ws["D1"] = "description"
    ws["E1"] = "notation"
    ws["F1"] = "note"
    ws["G1"] = "source"
    ws["H1"] = "broader"
    listOfKeys = list(data.keys())
    print(data.keys())
    counter = 2
    for i in range(0,len(listOfKeys)):
        if listOfKeys[i] is not "":
            ws["A"+str(counter)] = str(counter)
            print(listOfKeys[i])
            ws["B" + str(counter)] =  listOfKeys[i]
            #ws["C"+str(i)] = ""
            #ws["D"+str(i)] = ""
            #ws["E"+str(i)] = ""
            #ws["F"+str(i)] = ""
            ws["G" + str(counter)] = "http://mssanz.org.au/def/keyword/" + listOfKeys[i].replace(" ","_")
            #ws["H"+str(i)] = ""
            counter += 1
    wb.save(path)

putHere = defaultdict(list)
keyw = PdfExtractor.KeywordExtract()
putHere = keyw.fileToDictionary(putHere,"test.txt")
print(putHere)
dataset1 = putHere
#-take every filename and move keywords to new dictionary

tempdictionary = defaultdict(list)
listOfKeys = list(dataset1.keys())
print(listOfKeys)

for i in range (len(listOfKeys)):
    tempdictionary[fileLinkReverter(listOfKeys[i])] = dataset1[listOfKeys[i]]

with open("dataset1.txt","w") as file:
    file.write(json.dumps(tempdictionary))

#reformat dataset to id : label,description,notation,altLabel,note,source,broader,source

#-inverse dictionary
inverseDataset1 = defaultdict(list)
print(dataset1.items())
for key in dataset1.keys():
    for item in dataset1[key]:
        #print("key is " + key, "item is " + item)
        inverseDataset1[item].append(fileLinkReverter(key))
print(inverseDataset1)
print(len(inverseDataset1.keys()))
for item in inverseDataset1.keys():
    print(item)

writeToExcel("dataset2.xlsx" , inverseDataset1)
print("done")

#puts all the filenames to links
#for key in list(tempdictionary.keys()):
#    for item in tempdictionary[key]:
#        inverseDataset1[item].append(key)
print(inverseDataset1)
#for every single element in key, add key to the dataset2